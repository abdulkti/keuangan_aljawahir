import requests
import re
import time
import openpyxl

BASE_URL = "https://keuangan.aljawahirattarbawi.com"
EMAIL = "superadmin@aljawahir.sch.id"
PASSWORD = "superadmin123"
DELAY = 0.3
TANGGAL = "2026-06-30"
KETERANGAN = "THT Lampau"
EXCEL_PATH = "/Users/macbook/Downloads/THT_tabungan hari tua - rekap.xlsx"
EXCEL_SHEET = "2026-2027"

EXCEL_TO_APP = {
    'Faza Khairani': 28,
    'Elsa Anatasya': 46,
    'Uswatun Hasanah': 48,
    'Kharunnisa': 45,
    'Dewi Gustiani': 29,
    'Nur Zahara': 44,
    'Mika Nurmalaya': 39,
    'Arfikah Diah Mendasari': 40,
    'Hayati Purnaweni': 31,
    'Widya Ningsih': 41,
    'Sutriani': 32,
    'Miftah Pratiwi': 47,
    'Chairani': 30,
    'Nur Futri Utami': 35,
    'Rusdan': 34,
    'Mila Pratiwi': 42,
    'Rodia Insani Harahap': 36,
    'Sufi El Farobi': 54,
    'Ahmad Fathulillah': 52,
    'Ramadinisa': 53,
    'Ervina Rahma Safira': 49,
    'Kak Murni': 37,
    'Junaidi': 33,
    'Novalia Pratiwi': 2,
    'Indah Pertiwi': 3,
    'Erni Justika': 4,
    'Indah Nuraini': 5,
    'Rizka Ramadana': 6,
    'Rini Trinasya Audy': 7,
    'Khairunnisah': 1,
    'Adelia Tristanti': 8,
    'Putri Sakina Najwa': 9,
    'Ayu Rahayu': 10,
    'Nurainun': 11,
    'Khairunnisa, S.Tr.P': 12,
    'Icha Inggrid Lestary': 13,
    'Adelia, S.Pd': 15,
    'Fiki Hidayat': 17,
    'Annisaa Pratiwi Simanjuntak': 18,
    'Vivi Destri Yumielda': 19,
    'Shafira Anggita Purba': 22,
    'Adhe Eva Yolanda': 23,
    'Berby Yoreza': 24,
    'Nurliana Amelda': 25,
    'Dara Aisya': 27,
}

session = requests.Session()
csrf_token = None
csrf_form = None
created = 0
errors = 0


def login():
    global errors
    r = session.get(f"{BASE_URL}/login")
    m = re.search(r'name="csrf_test_name"\s+value="([^"]+)"', r.text)
    if not m:
        print("ERROR: Could not find csrf_test_name on login page")
        errors += 1
        return False
    csrf = m.group(1)
    r = session.post(f"{BASE_URL}/auth/login", data={
        "csrf_test_name": csrf,
        "email": EMAIL,
        "password": PASSWORD,
    }, allow_redirects=True)
    if "/dashboard" in r.url or r.status_code == 200:
        print("Login successful")
        return True
    print(f"ERROR: Login failed, status={r.status_code}, url={r.url}")
    errors += 1
    return False


def refresh_csrf():
    global csrf_token, csrf_form
    r = session.get(f"{BASE_URL}/tht")
    m1 = re.search(r'id="csrfToken"\s+value="([^"]+)"', r.text)
    m2 = re.search(r'name="csrf_test_name"\s+value="([^"]+)"', r.text)
    if m1:
        csrf_token = m1.group(1)
    if m2:
        csrf_form = m2.group(1)
    return bool(m1 and m2)


def ensure_session():
    global csrf_token, csrf_form
    r = session.get(f"{BASE_URL}/tht")
    if "/login" in r.url:
        print("  Session expired, re-logging in...")
        if not login():
            return False
        time.sleep(DELAY)
        return refresh_csrf()
    m1 = re.search(r'id="csrfToken"\s+value="([^"]+)"', r.text)
    m2 = re.search(r'name="csrf_test_name"\s+value="([^"]+)"', r.text)
    if m1:
        csrf_token = m1.group(1)
    if m2:
        csrf_form = m2.group(1)
    return bool(m1 and m2)


def check_session_valid():
    try:
        r = session.get(
            f"{BASE_URL}/tht/riwayat/1",
            headers={"X-Requested-With": "XMLHttpRequest"}
        )
        r.json()
        return True
    except Exception:
        return False


def read_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[EXCEL_SHEET]
    entries = []
    for row in range(6, 55):
        name_cell = ws.cell(row=row, column=2).value
        lampau_cell = ws.cell(row=row, column=3).value
        if not name_cell:
            continue
        name_cell = str(name_cell).strip()
        # Parse lampau amount
        if lampau_cell is None or str(lampau_cell).strip() in ('', '-', '0', 'None'):
            lampau = 0
        else:
            try:
                lampau = int(float(str(lampau_cell).replace(',', '').strip()))
            except (ValueError, TypeError):
                lampau = 0
        # Skip if lampau is 0
        if lampau <= 0:
            print(f"  SKIP (lampau=0): {name_cell}")
            continue
        # Match name to app guru ID using case-insensitive partial matching
        matched_key = None
        name_lower = name_cell.lower()
        for key in EXCEL_TO_APP:
            key_lower = key.lower()
            if key_lower in name_lower or name_lower in key_lower:
                matched_key = key
                break
        if matched_key is None:
            print(f"  SKIP (no match): {name_cell}")
            continue
        guru_id = EXCEL_TO_APP[matched_key]
        entries.append((guru_id, matched_key, lampau))
    return entries


def create_tht_lampau(entries):
    global created, errors
    print(f"\n=== Creating THT Lampau for {len(entries)} gurus ===")
    if not refresh_csrf():
        print("ERROR: Could not get CSRF")
        return

    total = len(entries)
    ops_since_refresh = 0
    for i, (guru_id, name, lampau) in enumerate(entries):
        if ops_since_refresh >= 15:
            if not check_session_valid():
                if not ensure_session():
                    print("  FATAL: Cannot re-establish session")
                    return
            else:
                refresh_csrf()
            ops_since_refresh = 0

        print(f"  [{i+1}/{total}] Creating THT Lampau for guru_id={guru_id} ({name}): Rp {lampau}")
        try:
            r = session.post(
                f"{BASE_URL}/tht/setor",
                data={
                    "csrf_test_name": csrf_form,
                    "guru_id": guru_id,
                    "jumlah": lampau,
                    "tanggal": TANGGAL,
                    "keterangan": KETERANGAN,
                },
                headers={
                    "X-Requested-With": "XMLHttpRequest",
                    "X-CSRF-TOKEN": csrf_token,
                },
                allow_redirects=False,
            )
            ops_since_refresh += 1
            if r.status_code in (200, 302, 303):
                created += 1
            else:
                if r.status_code in (403,) or "login" in r.text[:500].lower():
                    print(f"    Session issue, re-login...")
                    if ensure_session():
                        r = session.post(
                            f"{BASE_URL}/tht/setor",
                            data={
                                "csrf_test_name": csrf_form,
                                "guru_id": guru_id,
                                "jumlah": lampau,
                                "tanggal": TANGGAL,
                                "keterangan": KETERANGAN,
                            },
                            headers={
                                "X-Requested-With": "XMLHttpRequest",
                                "X-CSRF-TOKEN": csrf_token,
                            },
                            allow_redirects=False,
                        )
                        ops_since_refresh = 1
                        if r.status_code in (200, 302, 303):
                            created += 1
                        else:
                            print(f"    ERROR status={r.status_code} body={r.text[:200]}")
                            errors += 1
                    else:
                        print(f"    FATAL: cannot re-login")
                        errors += 1
                else:
                    print(f"    ERROR status={r.status_code} body={r.text[:200]}")
                    errors += 1
        except Exception as e:
            print(f"    EXCEPTION: {e}")
            if not check_session_valid():
                ensure_session()
            errors += 1
        time.sleep(DELAY)


def verify_guru_28():
    print("\n=== Verification: Guru 28 (Faza Khairani) riwayat ===")
    try:
        r = session.get(
            f"{BASE_URL}/tht/riwayat/28",
            headers={"X-Requested-With": "XMLHttpRequest"}
        )
        data = r.json()
        html = data.get("html", "")
        has_lampau = "THT Lampau" in html
        has_juli = "Iuran Juli" in html
        print(f"  THT Lampau present: {has_lampau}")
        print(f"  Iuran Juli present: {has_juli}")
    except Exception as e:
        print(f"  Verification error: {e}")


def main():
    print("=== Reading Excel data ===")
    entries = read_excel()
    print(f"  Found {len(entries)} entries with lampau > 0\n")

    if not login():
        return
    time.sleep(DELAY)

    create_tht_lampau(entries)
    time.sleep(DELAY)

    verify_guru_28()

    print(f"\nCREATED: {created} THT Lampau transactions")
    if errors:
        print(f"ERRORS: {errors}")


if __name__ == "__main__":
    main()
